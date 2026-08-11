"""
Mojave HVAC — Long Lead Forecast Builder
=========================================
Reads source files from data/ and injects updated DASH data
into output/index.html.

Usage:
    python build.py

Source files (place in data/ folder, most recent wins):
    Items   : ToExcel_Items*.csv, items*.csv, Long lead items*.csv
    CRM     : CRM demand planning tool*.xlsx
    Planning: ToExcel_PlanningDetail*.csv, planning detail*.csv
"""

import csv, io, json, os, glob, re, shutil, sys
from datetime import datetime

# ── Config ────────────────────────────────────────────────────────────────────

TODAY      = datetime.today()
SHOW_IDX   = [0, 1, 2, 3, 4]
OBS_OVERRIDE = {'P-003005', 'P-003006'}

COMM_MAP = {
    'Compressor':          ['compressor'],
    'Fans':                ['fans', 'fan'],
    'Evaporator Coils':    ['evaporator coil', 'evap co'],
    'Micro Channel Coils': ['micro channel', 'xcond', 'regen coil'],
    'Liquid Desiccant':    ['liquid desiccant', 'salt'],
    'VFD':                 ['vfd'],
}

MANUAL_OVERRIDES = {
    'P-002488': 7.0,
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR   = os.path.join(SCRIPT_DIR, 'data')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')
HTML_FILE  = os.path.join(OUTPUT_DIR, 'index.html')

# ── Helpers ───────────────────────────────────────────────────────────────────

def to_float(v):
    try:    return float(str(v).replace(',', '').strip()) if v is not None else 0.0
    except: return 0.0

def parse_date(s):
    for fmt in ['%m/%d/%Y', '%Y-%m-%d']:
        try: return datetime.strptime(str(s).strip(), fmt)
        except: pass
    return None

def read_tsv(path):
    with open(path, 'rb') as f: raw = f.read()
    text = raw.decode('utf-16') if raw[:2] == b'\xff\xfe' else raw.decode('utf-8-sig')
    return list(csv.DictReader(io.StringIO(text), delimiter='\t'))

def norm_comm(desc):
    d = str(desc).strip().lower()
    for comm, aliases in COMM_MAP.items():
        for a in aliases:
            if a in d: return comm
    return None

def latest_file(pattern):
    matches = glob.glob(pattern)
    return max(matches, key=os.path.getmtime) if matches else None

# ── Find source files ─────────────────────────────────────────────────────────

def find_items():
    for pat in ['ToExcel_Items*.csv', 'items*.csv', 'Items*.csv',
                'Long lead items*.csv', 'long lead items*.csv']:
        f = latest_file(os.path.join(DATA_DIR, pat))
        if f: return f
    return None

def find_crm():
    f = latest_file(os.path.join(DATA_DIR, 'CRM demand planning tool*.xlsx'))
    return f or latest_file(os.path.join(DATA_DIR, 'CRM*.xlsx'))

def find_planning():
    for pat in ['ToExcel_PlanningDetail*.csv', 'planning detail*.csv',
                'Planning detail*.csv', 'ToExcel_Planning*.csv']:
        f = latest_file(os.path.join(DATA_DIR, pat))
        if f: return f
    return None

# ── Load items ────────────────────────────────────────────────────────────────

def load_items(path):
    print(f"  Items : {os.path.basename(path)}")
    if path.lower().endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
        rows = [{h: row[i] for i,h in enumerate(headers) if h}
                for row in ws.iter_rows(min_row=2, values_only=True)]
    else:
        rows = read_tsv(path)

    item_rows = {str(r.get('Item') or '').strip(): r
                 for r in rows if str(r.get('Item') or '').strip()}

    eligible = []
    for item, r in item_rows.items():
        desc = str(r.get('Description') or '')
        lt   = to_float(r.get('Lead Time'))
        if lt <= 59: continue
        if desc.upper().startswith('OBS') and item not in OBS_OVERRIDE: continue
        comm_desc = str(r.get('Commodity Description') or '').strip()
        eligible.append({
            'item': item, 'desc': desc,
            'commodity': norm_comm(comm_desc) or 'Missing Commodity',
            'raw_commodity': comm_desc, 'lt': lt,
            'qoh':   to_float(r.get('Quantity On Hand')),
            'qord':  to_float(r.get('Quantity Ordered')),
            'alloc': to_float(r.get('Allocated To Prod')),
            'ss':    to_float(r.get('Safety Stock')),
            'forecast_values': [], 'sched_demand': 0.0,
        })
    print(f"    {len(eligible)} eligible parts (LT > 59 days)")
    return eligible

# ── Load CRM ──────────────────────────────────────────────────────────────────

def load_crm(path, eligible):
    print(f"  CRM   : {os.path.basename(path)}")
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Parts Forecast']
    periods = []
    for c in range(4, ws.max_column+1):
        val = ws.cell(row=3, column=c).value
        if val: periods.append(str(val).strip())
        else: break
    n = len(periods)
    print(f"    {n} periods starting {periods[0] if periods else '?'}")
    pFcst = {}; erpFcst = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        p_num = str(row[1]).strip() if row[1] else ''
        erp   = row[2]
        if p_num.startswith('P-'):
            if p_num not in pFcst: pFcst[p_num] = [0.0]*n
            for c in range(n):
                v = row[3+c]; pFcst[p_num][c] += float(v) if v else 0.0
        if erp and isinstance(erp, (int, float)):
            s = str(int(erp))
            if s not in erpFcst: erpFcst[s] = [0.0]*n
            for c in range(n):
                v = row[3+c]; erpFcst[s][c] += float(v) if v else 0.0
    for p in eligible:
        p['forecast_values'] = pFcst.get(p['item']) or erpFcst.get(p['item']) or [0.0]*n
    return periods

# ── Load planning ─────────────────────────────────────────────────────────────

def load_planning(path, eligible):
    print(f"  Plan  : {os.path.basename(path)}")
    outstd = {}
    for r in read_tsv(path):
        item = str(r.get('Item') or '').strip()
        req  = to_float(r.get('Outstanding Requirement'))
        if req <= 0: continue
        if item.startswith('P-'):
            outstd[item] = outstd.get(item, 0.0) + req
        else:
            d = parse_date(r.get('Date', ''))
            if d and d >= TODAY: outstd[item] = outstd.get(item, 0.0) + req
    outstd.update(MANUAL_OVERRIDES)
    for p in eligible: p['sched_demand'] = outstd.get(p['item'], 0.0)
    print(f"    {sum(1 for p in eligible if p['sched_demand'] > 0)} parts with scheduled demand")

# ── Build & inject ────────────────────────────────────────────────────────────

def build_and_inject(periods, eligible):
    parts_js = []
    for p in eligible:
        fv = '[' + ','.join(str(v) for v in p['forecast_values']) + ']'
        parts_js.append(
            '{item:'  + json.dumps(p['item'])  + ',desc:' + json.dumps(p['desc']) +
            ',commodity:' + json.dumps(p['commodity']) +
            ',raw_commodity:' + json.dumps(p.get('raw_commodity','')) +
            ',lt:' + str(p['lt']) + ',qoh:' + str(p['qoh']) + ',qord:' + str(p['qord']) +
            ',alloc:' + str(p.get('alloc',0)) + ',ss:' + str(p.get('ss',0)) +
            ',forecast_values:' + fv + ',sched_demand:' + str(p['sched_demand']) + '}'
        )
    periods_js = '[' + ','.join(json.dumps(x) for x in periods) + ']'
    dash_line  = ('var DASH = { periods:' + periods_js +
                  ', showIdx:[0,1,2,3,4],\nparts:[' + ','.join(parts_js) + ']};')

    with open(HTML_FILE, 'r', encoding='utf-8') as f: lines = f.readlines()
    start = next(i for i,l in enumerate(lines) if l.startswith('var DASH = {'))
    end   = start + 1
    if lines[end].startswith('parts:['): end += 1
    lines[start:end] = [dash_line + '\n']
    with open(HTML_FILE, 'w', encoding='utf-8') as f: f.writelines(lines)
    print(f"  Output: output/index.html  ({os.path.getsize(HTML_FILE):,} bytes)")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"\nMojave Long Lead Forecast — {TODAY.strftime('%Y-%m-%d')}\n")
    items_path    = find_items()
    crm_path      = find_crm()
    planning_path = find_planning()
    missing = [n for n,p in [('Items',items_path),('CRM',crm_path),('Planning',planning_path)] if not p]
    if missing:
        print(f"ERROR: Missing source files: {', '.join(missing)}")
        print(f"Place them in: {DATA_DIR}")
        sys.exit(1)

    print("Loading...")
    eligible = load_items(items_path)
    periods  = load_crm(crm_path, eligible)
    load_planning(planning_path, eligible)

    print("\nBuilding...")
    build_and_inject(periods, eligible)

    from collections import Counter
    counts = Counter(p['commodity'] for p in eligible)
    print("\nParts by commodity:")
    for c in ['Compressor','Fans','Evaporator Coils','Micro Channel Coils',
              'Liquid Desiccant','VFD','Missing Commodity']:
        print(f"  {c}: {counts.get(c,0)}")
    print("\nDone. Open output/index.html in your browser.\n")

if __name__ == '__main__':
    main()
